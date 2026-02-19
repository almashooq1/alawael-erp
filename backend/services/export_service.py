"""
خدمة التصدير - Phase 3
Export Service - PDF, Excel, CSV
"""

from io import BytesIO, StringIO
import csv
import json
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """خدمة تصدير التقارير بصيغ متعددة"""
    
    def __init__(self):
        """تهيئة الخدمة"""
        self.formats = ['pdf', 'excel', 'csv', 'json']
    
    # ==========================================
    # 1. تصدير CSV
    # ==========================================
    
    def export_as_csv(self, report_data: dict, filename: str = None) -> BytesIO:
        """
        تصدير التقرير كملف CSV
        
        Args:
            report_data: بيانات التقرير
            filename: اسم الملف
        
        Returns:
            BytesIO: محتوى الملف
        """
        try:
            output = StringIO()
            
            # كتابة رأس التقرير
            report_type = report_data.get('report_type', 'Report')
            student_name = report_data.get('student_name', 'Unknown')
            generated_at = report_data.get('generated_at', datetime.now().isoformat())
            
            output.write(f"Report Type,{report_type}\n")
            output.write(f"Student,{student_name}\n")
            output.write(f"Generated At,{generated_at}\n")
            output.write("\n")
            
            # كتابة البيانات الرئيسية
            for key, value in report_data.items():
                if key not in ['report_type', 'student_name', 'generated_at', 'details']:
                    if isinstance(value, dict):
                        output.write(f"{key}\n")
                        for k, v in value.items():
                            output.write(f",{k},{v}\n")
                    elif isinstance(value, list):
                        output.write(f"{key}\n")
                        for item in value:
                            output.write(f",{item}\n")
                    else:
                        output.write(f"{key},{value}\n")
            
            # تحويل إلى BytesIO
            csv_bytes = BytesIO(output.getvalue().encode('utf-8-sig'))
            csv_bytes.seek(0)
            
            logger.info(f"CSV export created: {filename or 'report.csv'}")
            return csv_bytes
            
        except Exception as e:
            logger.error(f"Error creating CSV: {str(e)}")
            raise
    
    # ==========================================
    # 2. تصدير JSON
    # ==========================================
    
    def export_as_json(self, report_data: dict, filename: str = None) -> BytesIO:
        """
        تصدير التقرير كملف JSON
        
        Args:
            report_data: بيانات التقرير
            filename: اسم الملف
        
        Returns:
            BytesIO: محتوى الملف
        """
        try:
            # تحضير البيانات
            export_data = {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'export_format': 'json',
                    'version': '1.0'
                },
                'report': report_data
            }
            
            # تحويل إلى JSON مع دعم UTF-8
            json_string = json.dumps(export_data, ensure_ascii=False, indent=2)
            json_bytes = BytesIO(json_string.encode('utf-8'))
            json_bytes.seek(0)
            
            logger.info(f"JSON export created: {filename or 'report.json'}")
            return json_bytes
            
        except Exception as e:
            logger.error(f"Error creating JSON: {str(e)}")
            raise
    
    # ==========================================
    # 3. تصدير PDF (باستخدام reportlab)
    # ==========================================
    
    def export_as_pdf(self, report_data: dict, filename: str = None) -> BytesIO:
        """
        تصدير التقرير كملف PDF
        
        Args:
            report_data: بيانات التقرير
            filename: اسم الملف
        
        Returns:
            BytesIO: محتوى الملف
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            
            # إنشاء BytesIO لـ PDF
            pdf_buffer = BytesIO()
            
            # إنشاء المستند
            doc = SimpleDocTemplate(
                pdf_buffer,
                pagesize=A4,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            # إعدادات النمط
            styles = getSampleStyleSheet()
            
            # نمط عربي (إذا كان متاحاً)
            try:
                # محاولة استخدام خط عربي
                arabic_font = ParagraphStyle(
                    'Arabic',
                    parent=styles['Normal'],
                    fontSize=12,
                    textColor=colors.black
                )
            except:
                arabic_font = styles['Normal']
            
            # محتوى PDF
            story = []
            
            # العنوان
            title = Paragraph(
                f"<b>التقرير: {report_data.get('report_type', 'Report')}</b>",
                styles['Heading1']
            )
            story.append(title)
            story.append(Spacer(1, 0.3*inch))
            
            # معلومات الطالب
            student_name = report_data.get('student_name', 'Unknown')
            student_info = Paragraph(
                f"<b>اسم الطالب:</b> {student_name}",
                styles['Normal']
            )
            story.append(student_info)
            story.append(Spacer(1, 0.2*inch))
            
            # تاريخ التقرير
            generated_at = report_data.get('generated_at', datetime.now().isoformat())
            date_info = Paragraph(
                f"<b>تاريخ الإصدار:</b> {generated_at}",
                styles['Normal']
            )
            story.append(date_info)
            story.append(Spacer(1, 0.3*inch))
            
            # البيانات الرئيسية
            data_table = []
            for key, value in report_data.items():
                if key not in ['report_type', 'student_name', 'generated_at', 'details']:
                    if not isinstance(value, (dict, list)):
                        data_table.append([key, str(value)])
            
            if data_table:
                table = Table(data_table, colWidths=[2*inch, 4*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(table)
            
            # بناء PDF
            doc.build(story)
            
            # العودة إلى البداية
            pdf_buffer.seek(0)
            
            logger.info(f"PDF export created: {filename or 'report.pdf'}")
            return pdf_buffer
            
        except ImportError:
            logger.warning("reportlab not installed, falling back to text representation")
            return self._export_as_text_pdf(report_data, filename)
        except Exception as e:
            logger.error(f"Error creating PDF: {str(e)}")
            raise
    
    # ==========================================
    # 4. تصدير Excel (باستخدام openpyxl)
    # ==========================================
    
    def export_as_excel(self, report_data: dict, filename: str = None) -> BytesIO:
        """
        تصدير التقرير كملف Excel
        
        Args:
            report_data: بيانات التقرير
            filename: اسم الملف
        
        Returns:
            BytesIO: محتوى الملف
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # إنشاء Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            
            # تنسيق الأعمدة
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40
            
            # تنسيق الخلايا
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # رقم الصف
            row = 1
            
            # العنوان
            title = f"التقرير: {report_data.get('report_type', 'Report')}"
            ws[f'A{row}'] = title
            ws[f'A{row}'].font = Font(bold=True, size=14)
            ws.merge_cells(f'A{row}:B{row}')
            row += 2
            
            # معلومات الطالب
            ws[f'A{row}'] = "اسم الطالب"
            ws[f'B{row}'] = report_data.get('student_name', 'Unknown')
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            # تاريخ الإصدار
            ws[f'A{row}'] = "تاريخ الإصدار"
            ws[f'B{row}'] = report_data.get('generated_at', datetime.now().isoformat())
            ws[f'A{row}'].font = Font(bold=True)
            row += 2
            
            # رأس الجدول
            ws[f'A{row}'] = "الحقل"
            ws[f'B{row}'] = "القيمة"
            ws[f'A{row}'].fill = header_fill
            ws[f'B{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'].font = header_font
            row += 1
            
            # البيانات
            for key, value in report_data.items():
                if key not in ['report_type', 'student_name', 'generated_at', 'details']:
                    ws[f'A{row}'] = str(key)
                    if isinstance(value, (dict, list)):
                        ws[f'B{row}'] = json.dumps(value, ensure_ascii=False, indent=2)
                    else:
                        ws[f'B{row}'] = str(value)
                    
                    ws[f'A{row}'].border = border
                    ws[f'B{row}'].border = border
                    ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                    ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
                    row += 1
            
            # حفظ الملف في BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logger.info(f"Excel export created: {filename or 'report.xlsx'}")
            return excel_buffer
            
        except ImportError:
            logger.warning("openpyxl not installed, falling back to CSV")
            return self.export_as_csv(report_data, filename)
        except Exception as e:
            logger.error(f"Error creating Excel: {str(e)}")
            raise
    
    # ==========================================
    # 5. دالة مساعدة - تصدير نص (عندما لا تكون reportlab متوفرة)
    # ==========================================
    
    def _export_as_text_pdf(self, report_data: dict, filename: str = None) -> BytesIO:
        """
        تصدير كـ text PDF (بديل عندما لا تكون reportlab متوفرة)
        """
        output = StringIO()
        output.write("="*60 + "\n")
        output.write(f"التقرير: {report_data.get('report_type', 'Report')}\n")
        output.write("="*60 + "\n\n")
        
        output.write(f"اسم الطالب: {report_data.get('student_name', 'Unknown')}\n")
        output.write(f"تاريخ الإصدار: {report_data.get('generated_at', datetime.now().isoformat())}\n\n")
        
        for key, value in report_data.items():
            if key not in ['report_type', 'student_name', 'generated_at']:
                output.write(f"{key}: {json.dumps(value, ensure_ascii=False)}\n")
        
        output.write("\n" + "="*60 + "\n")
        
        # تحويل إلى BytesIO
        text_bytes = BytesIO(output.getvalue().encode('utf-8'))
        text_bytes.seek(0)
        
        logger.info(f"Text PDF export created: {filename or 'report.txt'}")
        return text_bytes
    
    # ==========================================
    # 6. تصدير متعدد الصيغ
    # ==========================================
    
    def export_multi_format(self, report_data: dict, formats: list = None) -> dict:
        """
        تصدير التقرير بصيغ متعددة
        
        Args:
            report_data: بيانات التقرير
            formats: قائمة الصيغ المطلوبة
        
        Returns:
            dict: قاموس بالملفات المُصدّرة
        """
        if formats is None:
            formats = self.formats
        
        results = {}
        
        for fmt in formats:
            try:
                if fmt.lower() == 'pdf':
                    results['pdf'] = self.export_as_pdf(report_data)
                elif fmt.lower() == 'excel':
                    results['excel'] = self.export_as_excel(report_data)
                elif fmt.lower() == 'csv':
                    results['csv'] = self.export_as_csv(report_data)
                elif fmt.lower() == 'json':
                    results['json'] = self.export_as_json(report_data)
            except Exception as e:
                logger.error(f"Error exporting to {fmt}: {str(e)}")
                results[fmt] = None
        
        return results
    
    # ==========================================
    # 7. الحصول على نوع MIME المناسب
    # ==========================================
    
    @staticmethod
    def get_mime_type(file_format: str) -> str:
        """الحصول على نوع MIME للصيغة"""
        mime_types = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv',
            'json': 'application/json',
            'txt': 'text/plain',
        }
        return mime_types.get(file_format.lower(), 'application/octet-stream')
    
    @staticmethod
    def get_file_extension(file_format: str) -> str:
        """الحصول على امتداد الملف للصيغة"""
        extensions = {
            'pdf': 'pdf',
            'excel': 'xlsx',
            'csv': 'csv',
            'json': 'json',
            'txt': 'txt',
        }
        return extensions.get(file_format.lower(), 'bin')
